import openpyxl as xl
import urllib.parse
import os
import csv
import json
from glob import glob
import re
import chardet


def sanitize_with_only_chars(name, replace_char='_'):
    return re.sub(r"[^a-zA-Z0-9]", replace_char, name)


def sanitize_bucket_name(name, replace_char='_'):
    return re.sub(r"[^a-zA-Z0-9.\-_]", replace_char, name)


def sanitize_name(name, replace_char='_'):
    return re.sub(r'[^\w_.-]', replace_char, name)


def dump_xl_sheet(ws, rows, headers=None):
    if len(rows) == 0:
        return
    if headers is None:
        headers = rows[0].keys()

    for idx, header in enumerate(headers):
        ws.cell(row=1, column=idx+1, value=header)

    for row_idx, row in enumerate(rows):
        for col_idx, header in enumerate(headers):
            ws.cell(row=row_idx + 2, column=col_idx+1, value=row.get(header, None))


def change_extension(filepath, to_ext):
    if not to_ext.startswith('.'):
        to_ext = '.' + to_ext
    return os.path.splitext(filepath)[0] + to_ext


def dump_xl(filename, rows, headers=None):
    wb = xl.Workbook()
    ws = wb.active
    dump_xl_sheet(ws, rows, headers)
    wb.save(filename)


def read_xl(filename):
    wb = xl.load_workbook(filename)
    ws = wb.active
    values = []
    for row in ws:
        cell = row[0]
        values.append(cell.value)
    return values


def get_basename_wo_suffix(filename):
    return os.path.splitext(os.path.basename(filename))[0]


def from_browser_url_to_bucket_prefix(url):
    pi = urllib.parse.urlparse(url)
    bucket = os.path.basename(pi.path)
    query = urllib.parse.parse_qs(pi.query)
    prefix = query['prefix'][0]
    return bucket, prefix

def from_s3_url_to_bucket_prefix(url):
    pi = urllib.parse.urlparse(url)
    bucket = pi.hostname
    prefix = pi.path[1:]
    return bucket, prefix


def get_bucket_prefix(url):
    if url.startswith('s3'):
        return from_s3_url_to_bucket_prefix(url)
    else:
        return from_browser_url_to_bucket_prefix(url)


def read_csv(filename):
    with open(filename,encoding='utf8') as fp:
        return list(csv.DictReader(fp))


def read_csv_dict(filename):
    with open(filename, encoding='utf8') as fp:
        reader = csv.DictReader(fp)
        return list(reader)


def read_lines(filename):
    with open(filename) as fp:
        return [line.strip() for line in fp]


def dump_json(filename, loaded_data,ensure_ascii):
    with open(filename, 'w') as json_file:
        json.dump(loaded_data, json_file, indent=4, ensure_ascii=ensure_ascii)


def read_json(filename, encoding='utf8'):
    """
    read json file from local disk
    filename - name of the json file
    encoding - encoding with which to open the json file, default=utf8
    """
    with open(filename, encoding=encoding) as fp:
        return json.load(fp)


def dump_csv(filename, rows, headers=None, encoding='utf8'):
    """
    dump the rows of dictionaries into csv file
    filename - name of the csv file
    rows - list of dictionaries
    headers - sequence of headers if wanted to crop the header and give sequence
    encoding - encoding with which to write the file, default=utf8
    """
    with open(filename, 'w', newline='', encoding=encoding) as fp:
        if len(rows) == 0:
            return
        if headers is None:
            headers = rows[0].keys()
        writer = csv.DictWriter(fp, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def get_files_with_ext(directory, ext):
    """
    return all then files with given extensions from the directory
    directory - path of the directory
    ext - extension to search for e.g. .txt, .wav, .png
    """
    return list(glob(f'{directory}/**/*{ext}'), recursive=True)


def read_sc2_csv(filename):
    with open(filename) as fp:
        lines = fp.readlines()
    headers = lines[0].split(';')
    values = [l.split(';') for l in lines[1:]]
    return [{h:v for h,v in zip(headers,row)} for row in values]


def dump_csv_plain(filename, rows):
    with open(filename, 'w') as fp:
        writer = csv.writer(fp)
        for row in rows:
            writer.writerow(row)
